# PROCESSADOR FINAL CORRIGIDO - PRODUCAO DE CAMAS COM JANELA GRÁFICA
import zipfile
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from io import BytesIO
import unicodedata
from tkinter import Tk, filedialog, messagebox

def main():
    # Inicializa janela
    root = Tk()
    root.withdraw()
    messagebox.showinfo("Selecionar Arquivo ZIP", "Selecione o arquivo ZIP contendo as planilhas de produção")

    # Seleciona arquivo ZIP
    zip_path = filedialog.askopenfilename(
        title="Selecione o arquivo ZIP",
        filetypes=[("Arquivos ZIP", "*.zip")]
    )

    if not zip_path:
        messagebox.showerror("Erro", "Nenhum arquivo selecionado.")
        return

    # Nome do arquivo de saída com contador de arquivos no nome
    data_hoje = datetime.now().strftime("%d-%m-%y")

    messagebox.showinfo(
        "Selecionar Pasta de Saída",
        "Escolha a pasta onde o relatório será salvo",
    )
    output_dir = filedialog.askdirectory(title="Selecione a pasta de destino")

    if not output_dir:
        messagebox.showerror("Erro", "Nenhuma pasta selecionada.")
        return

    os.makedirs(output_dir, exist_ok=True)

    # Cabeçalhos padrão das planilhas
    colunas = ["#TAMANHO_DA_CAMA", "#TIPO_DE_TECIDO", "#COR_DO_TECIDO", "#QUANTIDADE_DE_PRODUCAO", "#DATA_DE_PRODUCAO"]

    # Funções auxiliares
    def limpar_valores(valor):
        if isinstance(valor, str):
            numeros = re.findall(r'\d+', valor)
            return int(numeros[0]) if numeros else None
        return int(valor) if pd.notnull(valor) else None

    def limpar_texto(texto):
        if isinstance(texto, str):
            texto = texto.upper()
            texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
            texto = re.sub(r'[^A-Z0-9 ]', '', texto)
            return texto.strip()
        return texto

    # DataFrames de saída
    df_box = pd.DataFrame(columns=colunas)
    df_bau = pd.DataFrame(columns=colunas)
    arquivos_processados = 0

    # Processa arquivos no ZIP
    with zipfile.ZipFile(zip_path, 'r') as z:
        for file_info in z.infolist():
            if file_info.filename.endswith(('.xlsx', '.xls')):
                arquivos_processados += 1
                nome_arquivo = os.path.basename(file_info.filename)
                data_arquivo = datetime(*file_info.date_time)
                data_nome = re.findall(r'(\d{2})[-_](\d{2})', nome_arquivo)
                if not data_nome:
                    continue
                dia, mes = data_nome[0]
                data_formatada = f"{dia}-{mes}-{str(data_arquivo.year)[-2:]}"

                with z.open(file_info) as f:
                    wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                    for sheet in wb.worksheets:
                        modo = None
                        box_data, bau_data = [], []
                        for row in sheet.iter_rows(values_only=True):
                            row_str = [str(cell).strip() if cell else "" for cell in row]
                            row_joined = " ".join(row_str).upper()
                            if "PROGRAMAÇÃO BOX" in row_joined:
                                modo = "BOX"
                                continue
                            elif "PROGRAMAÇÃO BAU" in row_joined:
                                modo = "BAU"
                                continue
                            elif modo == "BOX":
                                box_data.append(row_str + [data_formatada])
                            elif modo == "BAU":
                                bau_data.append(row_str + [data_formatada])

                        for dataset, df_target in [(box_data, df_box), (bau_data, df_bau)]:
                            if not dataset:
                                continue
                            df_temp = pd.DataFrame([d[:4] + [d[-1]] for d in dataset if len(d) >= 4], columns=colunas)
                            df_temp["#QUANTIDADE_DE_PRODUCAO"] = df_temp["#QUANTIDADE_DE_PRODUCAO"].apply(limpar_valores)
                            df_temp = df_temp[pd.to_numeric(df_temp["#QUANTIDADE_DE_PRODUCAO"], errors="coerce").notnull()]
                            df_temp["#QUANTIDADE_DE_PRODUCAO"] = df_temp["#QUANTIDADE_DE_PRODUCAO"].astype(int)
                            df_temp["#TIPO_DE_TECIDO"] = df_temp["#TIPO_DE_TECIDO"].apply(limpar_texto)
                            df_temp["#COR_DO_TECIDO"] = df_temp["#COR_DO_TECIDO"].apply(limpar_texto)
                            df_temp["#TAMANHO_DA_CAMA"] = df_temp["#TAMANHO_DA_CAMA"].apply(limpar_texto)
                            df_temp["#TIPO_DE_TECIDO"].replace({"LINUM": "SUEDE", "KORINO": "CORINO", "LIHNAO": "LINHAO"}, inplace=True)
                            df_temp = df_temp[df_temp["#TIPO_DE_TECIDO"].isin(["SUEDE", "LINHO", "LINHAO", "CORINO", "VELUDO"])]
                            if df_target is df_box:
                                df_box = pd.concat([df_box, df_temp], ignore_index=True)
                            else:
                                df_bau = pd.concat([df_bau, df_temp], ignore_index=True)

    # Agrupa e soma os dados finais
    for df in [df_box, df_bau]:
        df.dropna(subset=["#QUANTIDADE_DE_PRODUCAO"], inplace=True)
        df["#QUANTIDADE_DE_PRODUCAO"] = df["#QUANTIDADE_DE_PRODUCAO"].astype(int)

    df_box = df_box.groupby(["#TAMANHO_DA_CAMA", "#TIPO_DE_TECIDO", "#COR_DO_TECIDO", "#DATA_DE_PRODUCAO"], as_index=False).sum()
    df_bau = df_bau.groupby(["#TAMANHO_DA_CAMA", "#TIPO_DE_TECIDO", "#COR_DO_TECIDO", "#DATA_DE_PRODUCAO"], as_index=False).sum()

    nome_saida = os.path.join(output_dir, f"PRODUCAO_[{arquivos_processados}]_{data_hoje}.xlsx")

    # Exporta para Excel
    with pd.ExcelWriter(nome_saida, engine='openpyxl') as writer:
        df_box.to_excel(writer, sheet_name="BOX", index=False)
        df_bau.to_excel(writer, sheet_name="BAU", index=False)

    messagebox.showinfo("Sucesso", f"Relatório gerado na pasta:\n{output_dir}")
